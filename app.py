import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, time
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import re
from openpyxl.worksheet.table import Table, TableStyleInfo

st.title("Test")

recepcion = r"BD RECEPCION DE MATERIA PRIMA.xlsx"
enfriamiento = r"ENFRIAMIENTO 2025.xlsx"
volcado = r"BD VOLCADO DE MATERIA PRIMA.xlsx"



recpdf = pd.read_excel(recepcion,sheet_name="KF")



   
if recpdf["FECHA RECEPCION"].dtype == 'object':
    recpdf["FECHA RECEPCION"] = pd.to_datetime(recpdf["FECHA RECEPCION"], errors='coerce')
if recpdf["HORA RECEPCION"].dtype == 'object' or 'time' in str(recpdf["HORA RECEPCION"].dtype):
    recpdf["HORA RECEPCION"] = recpdf["HORA RECEPCION"].astype(str).str.extract(r'(\d{2}:\d{2}:\d{2})')[0]
elif pd.api.types.is_datetime64_any_dtype(recpdf["HORA RECEPCION"]):
    recpdf["HORA RECEPCION"] = recpdf["HORA RECEPCION"].dt.strftime('%H:%M:%S')

# Validar y corregir HORA RECEPCION para que siempre sea de la tarde
def corregir_hora_tarde(hora_str):
    if pd.isna(hora_str):
        return hora_str
    match = re.match(r"(\d{2}):(\d{2}):(\d{2})", str(hora_str))
    if not match:
        return hora_str
    h, m, s = map(int, match.groups())
    # Si la hora es menor a 12, sumamos 12 horas para que sea PM
    if h < 12:
        h += 12
        if h == 24:
            h = 12  # 12 PM
    return f"{h:02d}:{m:02d}:{s:02d}"

if "HORA RECEPCION" in recpdf.columns:
    recpdf["HORA RECEPCION"] = recpdf["HORA RECEPCION"].apply(corregir_hora_tarde)

# Now try the groupby
recpdf = recpdf.groupby(["FECHA RECEPCION","HORA RECEPCION","N° PALLET","CODIGO QR"])[["KILOS BRUTO"]].sum().reset_index()
#recpdf = recpdf.rename(columns={"KILOS NETO": "Cantidad Pallets"})
#FILTRO FOR COMS
recpdf = recpdf[recpdf["FECHA RECEPCION"] >= "2025-07-10"]
recpdf = recpdf.rename(columns={"CODIGO QR":"QR"})
recpdf = recpdf.drop(columns=["KILOS BRUTO"])
st.write(recpdf)
st.write(len(recpdf["QR"].unique()))
st.write(f"Recpcion: {recpdf.shape}")
#############################################################################
enfridf = pd.read_excel(enfriamiento,sheet_name="ENFRIAMIENTO")
print(enfridf.columns)
enfridf = enfridf.groupby(["FECHA","HORA INICIAL","QR","HORA FINAL"])[["FORMATO"]].count().reset_index()
enfridf = enfridf[enfridf["FECHA"] >= "2025-07-10"]
enfridf = enfridf.rename(columns={
    "FECHA": "FECHA ENFRIAMIENTO",
    "HORA FINAL": "HORA FINAL ENFRIAMIENTO",
    #"HORA INTERMEDIA": "HORA INTERMEDIA ENFRIAMIENTO",
    "HORA INICIAL": "HORA INICIAL ENFRIAMIENTO"
})
enfridf = enfridf.drop(columns=["FORMATO"])
st.write(enfridf.shape)
st.write(enfridf)
#################################################################################
volcadodf = pd.read_excel("BD VOLCADO DE MATERIA PRIMA.xlsx",sheet_name="BD")
volcadodf.to_excel("test.xlsx")
volcadodf["HORA FINAL "] = volcadodf["HORA FINAL "].fillna(pd.NaT)
if volcadodf["FECHA DE PROCESO"].dtype == 'object':
    volcadodf["FECHA DE PROCESO"] = pd.to_datetime(volcadodf["FECHA DE PROCESO"], errors='coerce')
if volcadodf["HORA INICIO"].dtype == 'object' or 'time' in str(volcadodf["HORA INICIO"].dtype):
    volcadodf["HORA INICIO"] = volcadodf["HORA INICIO"].astype(str).str.extract(r'(\d{2}:\d{2}:\d{2})')[0]
elif pd.api.types.is_datetime64_any_dtype(volcadodf["HORA INICIO"]):
    volcadodf["HORA INICIO"] = volcadodf["HORA INICIO"].dt.strftime('%H:%M:%S')
if volcadodf["HORA FINAL "].dtype == 'object' or 'time' in str(volcadodf["HORA FINAL "].dtype):
    volcadodf["HORA FINAL "] = volcadodf["HORA FINAL "].astype(str).str.extract(r'(\d{2}:\d{2}:\d{2})')[0]
elif pd.api.types.is_datetime64_any_dtype(volcadodf["HORA FINAL "]):
    volcadodf["HORA FINAL "] = volcadodf["HORA FINAL "].dt.strftime('%H:%M:%S')

    
volcadodf = volcadodf.groupby(["FECHA DE PROCESO","HORA INICIO","HORA FINAL ","QR","PROVEEDOR","FORMATO"])[["TIPO DE PRODUCTO"]].count().reset_index()
volcadodf = volcadodf[volcadodf["FECHA DE PROCESO"] >= "2025-07-10"]
volcadodf = volcadodf.rename(columns={"HORA INICIO": "HORA INICIO PROCESO","HORA FINAL ": "HORA FINAL PROCESO"})
volcadodf = volcadodf.drop(columns=["TIPO DE PRODUCTO"])
volcadodf.to_excel("test.xlsx")
st.write(volcadodf.shape)
st.dataframe(volcadodf)

dff = pd.merge(recpdf,enfridf,on="QR",how="left")
dff = pd.merge(dff,volcadodf,on="QR",how="left")
dff["FECHA RECEPCION"] = dff["FECHA RECEPCION"].dt.date
dff["FECHA ENFRIAMIENTO"] = dff["FECHA ENFRIAMIENTO"].dt.date
dff["FECHA DE PROCESO"] = dff["FECHA DE PROCESO"].dt.date


st.write(dff.shape)
st.write(dff)
#dff.to_excel("test.xlsx")

output_file = r"C:\Users\EdwardoGiampiereEnri\OneDrive - ALZA PERU GROUP S.A.C\PACKING TIEMPOS\TIEMPOS PACKING.xlsx"
#out_file_prueba = r"TIEMPOS PACKING TEST.xlsx"
with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    dff.to_excel(writer, index=False, sheet_name="TIEMPOS")
    ws = writer.sheets["TIEMPOS"]

    # Encabezados en negrita y fondo azul claro
    header_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
    for col_num, col in enumerate(dff.columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = Font(bold=True)
        cell.fill = header_fill

    # Ajustar ancho de columnas automáticamente
    for i, col in enumerate(dff.columns, 1):
        max_length = max(
            [len(str(cell.value)) if cell.value is not None else 0 for cell in ws[get_column_letter(i)]]
        )
        ws.column_dimensions[get_column_letter(i)].width = max_length + 2

    # Congelar la primera fila
    ws.freeze_panes = "A2"

    # Validar encabezados para tabla de Excel
    columnas_validas = True
    colnames = list(dff.columns)
    if any(pd.isna(col) or str(col).strip() == '' for col in colnames):
        columnas_validas = False
    if len(set(colnames)) != len(colnames):
        columnas_validas = False
    if any(any(c in str(col) for c in ['[', ']', '*', '?', '/', '\\']) for col in colnames):
        columnas_validas = False

    # Crear tabla de Excel real solo si los encabezados son válidos
    if columnas_validas:
        nrows = dff.shape[0] + 1  # +1 por encabezado
        ncols = dff.shape[1]
        last_col = get_column_letter(ncols)
        table_ref = f"A1:{last_col}{nrows}"
        tabla = Table(displayName="TIEMPOS_TABLA", ref=table_ref)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tabla.tableStyleInfo = style
        ws.add_table(tabla)
    else:
        st.warning("No se pudo crear la tabla de Excel porque los encabezados no son válidos para Excel. Solo se exportó el formato básico.")

st.success(f"Archivo Excel exportado con formato: {output_file}")




