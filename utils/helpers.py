import re
import pandas as pd
import openpyxl
import numpy as np
import calendar
from datetime import datetime
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import re
from openpyxl.worksheet.table import Table, TableStyleInfo
import io

change_month = {
    'Enero':1, 
    'Febrero':2, 
    'Marzo':3, 
    'Abril':4, 
    'Mayo':5, 
    'Jun':6,  
    'Jul':7, 
    'Ago':8, 
    'Set':9, 
    'Oct':10,
    'Nov':11, 
    'Dic':12
}
def limpiar_kg_exportables(valor):
        # Convertir a string para poder verificar si contiene punto
        valor_str = str(valor)
        
        if "." in valor_str:
            # Si contiene punto, convertir a float, multiplicar por 1000 y convertir a entero
            return float(valor_str) * 1000
        else:
            # Si no contiene punto, solo convertir a entero
            return int(valor_str)
        
def split_if_colon_at_3(texto):
    if isinstance(texto, str) and len(texto) > 2 and texto[2] == ':':
        return [texto[:2], texto[3:].strip()]
    else:
        return [None, texto]
    
def get_month_name(month_number: int) -> str:
    
    if not 1 <= month_number <= 12:
        raise ValueError("El número de mes debe estar entre 1 y 12")
        
    months = {
        1: "ENERO",
        2: "FEBRERO",
        3: "MARZO",
        4: "ABRIL",
        5: "MAYO",
        6: "JUNIO",
        7: "JULIO",
        8: "AGOSTO",
        9: "SEPTIEMBRE",
        10: "OCTUBRE",
        11: "NOVIEMBRE",
        12: "DICIEMBRE"
    }
    
    return months[month_number]
# Validar y corregir HORA RECEPCION para que siempre sea de la tarde
def corregir_hora_tarde(hora_str):
        if pd.isna(hora_str):
            return hora_str
        match = re.match(r"(\d{2}):(\d{2}):(\d{2})", str(hora_str))
        if not match:
            return hora_str
        h, m, s = map(int, match.groups())
        # Si la hora es menor a 12, sumamos 12 horas para que sea PM
        if h < 12:
            h += 12
            if h == 24:
                h = 12  # 12 PM
        return f"{h:02d}:{m:02d}:{s:02d}"


def create_format_excel(dff: pd.DataFrame, nombre_archivo: str) -> str:
     with pd.ExcelWriter(nombre_archivo, engine="openpyxl") as writer:
        dff.to_excel(writer, index=False, sheet_name="TIEMPOS")
        ws = writer.sheets["TIEMPOS"]

        # Encabezados en negrita y fondo azul claro
        header_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
        for col_num, col in enumerate(dff.columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True)
            cell.fill = header_fill

        # Ajustar ancho de columnas automáticamente
        for i, col in enumerate(dff.columns, 1):
            max_length = max(
                [len(str(cell.value)) if cell.value is not None else 0 for cell in ws[get_column_letter(i)]]
            )
            ws.column_dimensions[get_column_letter(i)].width = max_length + 2

        # Congelar la primera fila
        ws.freeze_panes = "A2"

        # Validar encabezados para tabla de Excel
        columnas_validas = True
        colnames = list(dff.columns)
        if any(pd.isna(col) or str(col).strip() == '' for col in colnames):
            columnas_validas = False
        if len(set(colnames)) != len(colnames):
            columnas_validas = False
        if any(any(c in str(col) for c in ['[', ']', '*', '?', '/', '\\']) for col in colnames):
            columnas_validas = False

        # Crear tabla de Excel real solo si los encabezados son válidos
        if columnas_validas:
            nrows = dff.shape[0] + 1  # +1 por encabezado
            ncols = dff.shape[1]
            last_col = get_column_letter(ncols)
            table_ref = f"A1:{last_col}{nrows}"
            tabla = Table(displayName="TIEMPOS_TABLA", ref=table_ref)
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            tabla.tableStyleInfo = style
            ws.add_table(tabla)
        else:
            print("No se pudo crear la tabla de Excel porque los encabezados no son válidos para Excel. Solo se exportó el formato básico.")

def create_format_excel_in_memory(dff: pd.DataFrame) -> bytes:
    """
    Crea un archivo Excel formateado en memoria y retorna los bytes
    
    Args:
        dff: DataFrame de pandas a formatear
    
    Returns:
        bytes: Contenido del archivo Excel formateado
    """
    # Crear buffer en memoria
    excel_buffer = io.BytesIO()
    
    with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
        dff.to_excel(writer, index=False, sheet_name="TIEMPOS")
        ws = writer.sheets["TIEMPOS"]

        # Encabezados en negrita y fondo azul claro
        header_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
        for col_num, col in enumerate(dff.columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True)
            cell.fill = header_fill

        # Ajustar ancho de columnas automáticamente
        for i, col in enumerate(dff.columns, 1):
            max_length = max(
                [len(str(cell.value)) if cell.value is not None else 0 for cell in ws[get_column_letter(i)]]
            )
            ws.column_dimensions[get_column_letter(i)].width = max_length + 2

        # Congelar la primera fila
        ws.freeze_panes = "A2"

        # Validar encabezados para tabla de Excel
        columnas_validas = True
        colnames = list(dff.columns)
        if any(pd.isna(col) or str(col).strip() == '' for col in colnames):
            columnas_validas = False
        if len(set(colnames)) != len(colnames):
            columnas_validas = False
        if any(any(c in str(col) for c in ['[', ']', '*', '?', '/', '\\']) for col in colnames):
            columnas_validas = False

        # Crear tabla de Excel real solo si los encabezados son válidos
        if columnas_validas:
            nrows = dff.shape[0] + 1  # +1 por encabezado
            ncols = dff.shape[1]
            last_col = get_column_letter(ncols)
            table_ref = f"A1:{last_col}{nrows}"
            tabla = Table(displayName="TIEMPOS_TABLA", ref=table_ref)
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            tabla.tableStyleInfo = style
            ws.add_table(tabla)
        else:
            print("⚠️  No se pudo crear la tabla de Excel porque los encabezados no son válidos. Solo se aplicó el formato básico.")
    
    # Obtener los bytes del archivo Excel formateado
    excel_data = excel_buffer.getvalue()
    excel_buffer.close()
    
    return excel_data

def get_download_url_by_name(json_data, name):
    """
    Busca en el JSON un archivo por su nombre y retorna su downloadUrl
    
    Args:
        json_data (list): Lista de diccionarios con información de archivos
        name (str): Nombre del archivo a buscar
    
    Returns:
        str: URL de descarga del archivo encontrado, o None si no se encuentra
    """
    for item in json_data:
        if item.get('name') == name:
            return item.get('@microsoft.graph.downloadUrl')
        
def structure_planilla_historica_like_estimate(df_planilla_historica):
    """
    Estructura la planilla histórica igual que estimate_current_planilla_by_previous:
    Para cada mes y proyecto, distribuye el costo total entre los días laborables (lunes a viernes) del mes.

    Args:
        df_planilla_historica (pd.DataFrame): DataFrame con columnas ['Mes', 'DESCRIPCION PROYECTO', 'Costos']

    Returns:
        pd.DataFrame: DataFrame con columnas ['DESCRIPCION PROYECTO', 'FECHA', 'TOTAL', 'MES_ACTUAL', 'AÑO_ACTUAL']
    """
    if not np.issubdtype(df_planilla_historica['Mes'].dtype, np.datetime64):
        df_planilla_historica = df_planilla_historica.copy()
        df_planilla_historica['Mes'] = pd.to_datetime(df_planilla_historica['Mes'])

    resultados = []
    for (año, mes), df_mes in df_planilla_historica.groupby([df_planilla_historica['Mes'].dt.year, df_planilla_historica['Mes'].dt.month]):
        # Solo días laborables (lunes=0, ..., viernes=4)
        dias_mes = pd.date_range(start=datetime(año, mes, 1), end=datetime(año, mes, pd.Period(f'{año}-{mes:02d}').days_in_month))
        dias_laborables = dias_mes[dias_mes.weekday < 5]  # lunes a viernes
        df_grouped = df_mes.groupby('DESCRIPCION PROYECTO', as_index=False)['Costos'].sum()
        for _, row in df_grouped.iterrows():
            costo_diario = row['Costos'] / len(dias_laborables) if len(dias_laborables) > 0 else 0
            for fecha in dias_laborables:
                resultados.append({
                    'DESCRIPCION PROYECTO': row['DESCRIPCION PROYECTO'],
                    'FECHA': fecha,
                    'TOTAL': costo_diario,
                })
    
    return pd.DataFrame(resultados)

def estimate_current_planilla_by_previous(df_planilla_historica):
    """
    Calcula la planilla "actual" (mes más reciente sin datos) usando la planilla del mes anterior,
    agrupando por proyecto y distribuyendo el costo total entre los días laborables (lunes a viernes) del mes actual.
    Si existe la planilla del mes actual, se usa esa. Si no, se usa la del mes anterior.

    Args:
        df_planilla_historica (pd.DataFrame): DataFrame con columnas ['Mes', 'DESCRIPCION PROYECTO', 'Costos']
            donde 'Mes' es tipo datetime o string 'YYYY-MM'.

    Returns:
        pd.DataFrame: DataFrame con columnas ['DESCRIPCION PROYECTO', 'DIA', 'COSTO_DIARIO', 'MES_ACTUAL', 'AÑO_ACTUAL']
        donde 'DIA' es una fecha completa (datetime)
    """
    # Normalizar columna 'Mes' a datetime
    if not np.issubdtype(df_planilla_historica['Mes'].dtype, np.datetime64):
        df_planilla_historica = df_planilla_historica.copy()
        df_planilla_historica['Mes'] = pd.to_datetime(df_planilla_historica['Mes'])

    # Encontrar el mes más reciente en la planilla histórica
    max_mes = df_planilla_historica['Mes'].max()
    
    año_max = max_mes.year
    mes_max = max_mes.month
    
    # Calcular el mes "actual" (el siguiente al más reciente en la planilla)
    if mes_max == 12:
        año_actual = año_max + 1
        mes_actual = 1
    else:
        año_actual = año_max
        mes_actual = mes_max + 1

    # Verificar si ya existe la planilla del mes actual
    existe_actual = (
        (df_planilla_historica['Mes'].dt.year == año_actual) &
        (df_planilla_historica['Mes'].dt.month == mes_actual)
    ).any()

    if existe_actual:
        # Usar la planilla del mes actual
        df_mes = df_planilla_historica[
            (df_planilla_historica['Mes'].dt.year == año_actual) &
            (df_planilla_historica['Mes'].dt.month == mes_actual)
        ]
        # Solo días laborables (lunes=0, ..., viernes=4)
        dias_mes_actual = pd.date_range(start=datetime(año_actual, mes_actual, 1), end=datetime(año_actual, mes_actual, pd.Period(f'{año_actual}-{mes_actual:02d}').days_in_month))
        dias_laborables = dias_mes_actual[dias_mes_actual.weekday < 5]
    else:
        # Usar la planilla del mes anterior
        df_mes = df_planilla_historica[
            (df_planilla_historica['Mes'].dt.year == año_max) &
            (df_planilla_historica['Mes'].dt.month == mes_max)
        ]
        # El mes actual es el siguiente al más reciente
        dias_mes_actual = pd.date_range(start=datetime(año_actual, mes_actual, 1), end=datetime(año_actual, mes_actual, pd.Period(f'{año_actual}-{mes_actual:02d}').days_in_month))
        dias_laborables = dias_mes_actual[dias_mes_actual.weekday < 5]

    # Agrupar por proyecto y sumar costos
    df_grouped = df_mes.groupby('DESCRIPCION PROYECTO', as_index=False)['Costos'].sum()
    num_dias_laborables = len(dias_laborables)
    df_grouped['COSTO_DIARIO'] = df_grouped['Costos'] / num_dias_laborables if num_dias_laborables > 0 else 0

    # Expandir a cada día laborable del mes actual como fecha completa
    df_result = pd.DataFrame([
        {
            'DESCRIPCION PROYECTO': row['DESCRIPCION PROYECTO'],
            'FECHA': fecha,
            'TOTAL': row['COSTO_DIARIO'],
            'MES_ACTUAL': mes_actual,
            'AÑO_ACTUAL': año_actual
        }
        for _, row in df_grouped.iterrows() for fecha in dias_laborables
    ])
    df_result = df_result[["DESCRIPCION PROYECTO","FECHA","TOTAL"]]
    return df_result


# Función para convertir fechas con formato mixto a YYYY/MM/DD
def convert_mixed_dates(date_series):
        def parse_date(date_str):
            if pd.isna(date_str) or date_str == '':
                return pd.NaT
            
            # Convertir a string y reemplazar guiones por barras
            date_str = str(date_str).replace("-", "/")
            
            # Si ya está en formato YYYY/MM/DD, mantenerlo
            if len(date_str.split('/')[0]) == 4:
                try:
                    return pd.to_datetime(date_str, format='%Y/%m/%d')
                except:
                    pass
            
            # Si está en formato DD/MM/YYYY, convertirlo
            try:
                return pd.to_datetime(date_str, format='%d/%m/%Y')
            except:
                return pd.NaT
        
        return date_series.apply(parse_date)

def transform_kg_text_rp_packing(text_num):
        if len(text_num) > 4 and (text_num[1] == "." or text_num[2] == "."):
            text_num = text_num.replace(".", "")
        else:
            text_num = text_num.replace(",", ".")
        return text_num.replace(",",".")